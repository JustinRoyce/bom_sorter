import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from pandas.core.accessor import register_series_accessor
import xlsxwriter

SCRIPT_PATH = __file__
SCRIPT_NAME = os.path.basename(SCRIPT_PATH)
ROOT_PATH = os.path.dirname(SCRIPT_PATH)
BOM_PATH = os.path.join(ROOT_PATH,"bom")
BOM_WB = os.path.join(BOM_PATH,"bom.xlsx")
CSV_HEADERS = ['Ref',
    'Qnty',
    'Value',
    'Cmp name',
    'Footprint',
    'Description',	
    'Manufacturer',	
    'MPN',	
    'Package',	
    'Rating',
    "PCB"]
CSV_HEADER_SIZE = 11

PASSIVE_C_CMP_NAME = "C"
PASSIVE_R_CMP_NAME = "R_US"

def append_BOM_data():
    frames = []
    csv_fpaths_list = []
    csv_names_list = []
    csv_files = os.listdir(BOM_PATH)
    df = pd.DataFrame()
 
    for csv_file in csv_files:
        extension = os.path.splitext(csv_file)[1][1:].strip() 
        isCSV_file = extension == "csv"
        if(isCSV_file): 
            csv_fpath = os.path.join(BOM_PATH,csv_file)
            csv_fpaths_list.append(csv_fpath)
            csv_name = csv_file.replace(".csv","")
            
            csv_names_list.append(csv_names_list)
            
            tmp = pd.read_csv(csv_fpath)
            tmp_shape = tmp.shape

            row = tmp_shape[0]
            col = tmp_shape[1]
            PCB_list = []
            for i in range(0,row):
                PCB_list.append(csv_name)

            #print("row:" + str(row))
            #print("col:" + str(col))
            tmp['PCB'] = PCB_list
            frames.append(tmp)
        
    result = pd.concat(frames)
    
    #remove garbage from headers
    frames_cols = list(result.columns.values)
    frames_col_len = len(frames_cols)
   
    if (frames_col_len !=  CSV_HEADER_SIZE):
        for n in range(CSV_HEADER_SIZE,frames_col_len,1):
            trash_header = frames_cols[n]
            #print(str(n)+":" + trash_header)
            result.pop(trash_header)

    return result


if __name__ == "__main__":
    
    BOM_Data = append_BOM_data()

    print("The column headers :")
    print(list(BOM_Data.columns.values))
    
     #PASSIVE COMPONENT NAMES:
    # Cmp name: C
    # Cmp name: R_US
    # Cmp name: LED
    
    
    indx = 0 
    jellybean_indx_list = []
    active_indx_list = []
    connectors_indx_list = []
    non_parts_indx_list = []
    
    refseries = BOM_Data["Ref"]
    cmp_series = BOM_Data['Cmp name']
    
    for refs in refseries:

        cmp = cmp_series.iloc[indx].strip()

        if(cmp == "C"):
            isJellybean = True
        elif(cmp == "R_US"):
            isJellybean = True
        elif(cmp == "LED"):        
            isJellybean = True
        elif(cmp == "LED_ALT"):        
            isJellybean = True    
        else: 
            isJellybean = False    


        ref_char_len = len(refs) 
        char_0 = refs[0]
        char_1 = refs[1]

        isTP_ref = char_0 == "T" and char_1 == "P"
        isMT_ref = char_0 == "M" and char_1 == "T"
        isJP_ref = char_0 == "J" and char_1 == "P"
        isMH_ref = char_0 == "M" and char_1 == "H"
        isH_ref = char_0 == "H" and char_1.isnumeric()

        isP_ref = char_0 == "P" and char_1.isnumeric()
        isJ_ref = char_0 == "J" and char_1.isnumeric()

        isConnector = isP_ref or isJ_ref 
        is_Non_parts = isJP_ref or isMT_ref or isTP_ref \
            or isMH_ref or isH_ref


        if(isJellybean):
            jellybean_indx_list.append(indx)

        if(isConnector):
            connectors_indx_list.append(indx)
            

        if(is_Non_parts):
            non_parts_indx_list.append(indx)


        active_logic = not isJellybean and not isConnector and \
                    not is_Non_parts 
        
        if(active_logic):     
            active_indx_list.append(indx)
        
        indx = indx + 1
    
    
    connectors_df = BOM_Data.iloc[connectors_indx_list]
    non_parts_df = BOM_Data.iloc[non_parts_indx_list]
    jellybean_df = BOM_Data.iloc[jellybean_indx_list]
    active_df = BOM_Data.iloc[active_indx_list]
 
    
    filepath = BOM_WB
    wb = openpyxl.Workbook()
    wb.save(filepath)

    
    
    book = load_workbook(BOM_WB)
    writer = pd.ExcelWriter(BOM_WB, engine='openpyxl')
    writer.book = book
    BOM_Data.to_excel(writer,sheet_name="raw_data",index = False,header= True)
    connectors_df.to_excel(writer,sheet_name="raw_connectors",index = False,header= True)
    jellybean_df.to_excel(writer,sheet_name="raw_jellybean_components",index = False,header= True)
    non_parts_df.to_excel(writer,sheet_name="raw_non_populated",index = False,header= True)
    active_df.to_excel(writer,sheet_name="raw_active_components",index = False,header= True)
    writer.save()
    
    #remove sheet called sheet
    wb = load_workbook(BOM_WB)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(BOM_WB)
    